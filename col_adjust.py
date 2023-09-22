import os
import openpyxl

for root, dirs, files in os.walk('./0919打乱/输出'):
    for f in files:
        Excel_path = os.path.join(root, f)
        workbook = openpyxl.load_workbook(Excel_path)

        sheetnames = workbook.sheetnames

        for sheet_name in sheetnames:
            sheet = workbook[sheet_name]
            for i, col in enumerate(sheet.columns):
                max_length = 0
                column = col[0].column_letter  # 获取列字母序号
                for cell in col:
                    try:
                        # 获取单元格中最长的文本长度
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 20  # 适配宽度多留较宽空隙
                sheet.column_dimensions[column].width = adjusted_width

        # 保存修改后的Excel文件
        workbook.save(Excel_path)

        # 关闭工作簿
        workbook.close()
