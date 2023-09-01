from openpyxl import load_workbook,Workbook
import os

root1 = "./data/按sheet拆分表格0719地市/"
os.makedirs(root1)
os.chdir(root1)

wb = load_workbook("D:/project/pandasD/data/tmp6.xlsx")

sheetnames = wb.sheetnames
print(sheetnames)

for name in sheetnames:
    ws = wb.get_sheet_by_name(name)
    print(ws)

# 创建新excel
    wb2 = Workbook()

# 获取当前sheet，调用正在运行的工作表格
    ws2 = wb2.active

# 两个for循环遍历整个excel的单元格内容
# enumerate() 函数用于将一个可遍历的数据对象(如列表、元组或字符串)组合为一个索引序列，同时列出数据和数据下标，一般用在 for 循环当中。

    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            # 写入新Excel
            ws2.cell(row=i+1, column=j+1, value=cell.value)
            # 设置新Sheet的名称
            ws2.title = name

    wb2.save(name + ".xlsx")






